from pathlib import Path
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from packages.companies.repository import create_company, find_duplicate_by_name
from packages.companies.schemas import CompanyCreate

EXPECTED_COLUMNS = {
    "name": {"company", "company name", "name", "empresa"},
    "domain": {"domain", "dominio"},
    "industry": {"industry", "sector", "industria"},
    "country": {"country", "pais", "país"},
    "employee_count": {"employees", "employee_count", "empleados"},
}


def import_companies_from_excel(
    session: Session,
    organization_id: UUID,
    file_path: Path,
    workspace_id: UUID | None = None,
) -> dict:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"created": 0, "duplicates": 0}

    headers = [_normalize_header(value) for value in rows[0]]
    mapping = _map_headers(headers)
    created = 0
    duplicates = 0

    for row in rows[1:]:
        raw = dict(zip(headers, row, strict=False))
        name = _get(raw, mapping, "name")
        if not name:
            continue
        if find_duplicate_by_name(session, organization_id, str(name)):
            duplicates += 1
            continue
        create_company(
            session,
            CompanyCreate(
                organization_id=organization_id,
                workspace_id=workspace_id,
                name=str(name),
                domain=_optional_str(_get(raw, mapping, "domain")),
                industry=_optional_str(_get(raw, mapping, "industry")),
                country=_optional_str(_get(raw, mapping, "country")),
                employee_count=_optional_int(_get(raw, mapping, "employee_count")),
                confidence=0.2,
            ),
        )
        created += 1

    return {"created": created, "duplicates": duplicates}


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def _map_headers(headers: list[str]) -> dict[str, str]:
    mapping = {}
    for field, aliases in EXPECTED_COLUMNS.items():
        for header in headers:
            if header in aliases:
                mapping[field] = header
                break
    return mapping


def _get(row: dict, mapping: dict[str, str], field: str) -> object:
    header = mapping.get(field)
    return row.get(header) if header else None


def _optional_str(value: object) -> str | None:
    return str(value).strip() if value not in (None, "") else None


def _optional_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    return int(value)
